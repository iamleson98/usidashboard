from fastapi import FastAPI
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import os
from datetime import datetime
import shutil

username = os.getlogin()
mealtime = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
abnormal = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
offwork = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
current_time = datetime.now().strftime("%m-%d-%Y-%H-%M-%S")
report_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
filename = current_time+"BreakTimeReport"
template_path = "template.xlsx"
output_path = f"C:/Users/{username}/OneDrive - Universal Global Scientific Industrial Corporation/USI_BreakTimeReport/{filename}.xlsx"

json_response = '''
[
    {"employee_id": 1, "in_time": "2025-04-23T08:00:00", "out_time": "2025-04-23T17:00:00", "minutes": 540, "break_type_id": "Abnormal"},
    {"employee_id": 2, "in_time": "2025-04-23T09:00:00", "out_time": "2025-04-23T18:00:00", "minutes": 480, "break_type_id": "Normal"}
]
'''

data = json.loads(json_response)
df = pd.DataFrame(data)
shutil.copyfile(template_path, output_path)
workbook = load_workbook(output_path)
sheet = workbook.active
sheet.cell(2,2,report_time)

for r_idx, row in df.iterrows():
    for c_idx, value in enumerate(row):
        sheet.cell(row=r_idx + 5, column=c_idx + 1, value=value)

for row in range(5, sheet.max_row + 1):
    cell = sheet.cell(row=row, column=4)
    if 'Abnormal' in sheet[f'E{row}'].value:
        cell.fill = abnormal
    if 'Mealtime' in sheet[f'E{row}'].value:
        cell.fill = mealtime
    if 'Offwork' in sheet[f'E{row}'].value:
        cell.fill = offwork

workbook.save(f'C:/Users/{username}/OneDrive - Universal Global Scientific Industrial Corporation/USI_BreakTimeReport/{filename}.xlsx')

